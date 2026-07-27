"""
GTI Quote Export Cleaner — deterministic cleaning pipeline.

Reads a legacy .xls "Quotes List" export (BIFF, via xlrd) and produces a
formatted .xlsx (via openpyxl). Pure Python: runs unchanged locally and in
the browser under Pyodide. No pandas, no network, no persistence.

The specification is PRD section 6. Every transformation here maps to a
numbered rule in that document and nothing else is applied to the data.

Entry point: clean_workbook(raw_bytes) -> dict
    {
      "ok": bool,
      "error": str | None,          # human-readable abort message when ok is False
      "xlsx": bytes | None,         # the cleaned workbook when ok is True
      "summary": { ... },           # counts + flags (see build_summary)
    }
"""

import io
import re
import datetime

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Fixed lookups (PRD 6.1, 6.3, 6.8). These are the ONLY lookups permitted.
# ---------------------------------------------------------------------------

# 6.1 furniture: rows whose joined text contains one of these are page furniture.
FURNITURE_SUBSTRINGS = ("QUOTES LIST", "FOR DATE RANGE")
# 6.1 residual sub-header fragments (the "No." / "By" line under the header row).
SUBHEADER_TOKENS = {"No.", "No", "by", "By"}

# 6.3 final schema, in exact output order, mapped to the export's own header labels.
# (final header -> source header label as it appears in the two-line header block)
FINAL_TO_SOURCE = [
    ("Account No.",  "Acct No."),
    ("Account Name", "Acct Name"),
    ("Quote",        "Est No."),
    ("Est. Date",    "Est Date"),
    ("Job Name",     "Order No"),
    ("Amount",       "Amount"),
    ("Invoice",      "Inv No."),
    ("GTI Comments", "Comments"),
    ("Product",      "Product"),
    ("SQFT",         "SQFT"),
    ("Created By",   "Created By"),
    ("Internal Note","Internal Note"),
]
SOURCE_COLUMNS = [f for f, _ in FINAL_TO_SOURCE]
REQUIRED_SOURCE_HEADERS = [s for _, s in FINAL_TO_SOURCE]

# 6.3b derived columns: computed from the GTI Comments text, not read from any
# source header. REQUIRED_SOURCE_HEADERS stays source-only because it drives the
# "is this the right report?" validation gate — dropping GTI Comments from the
# OUTPUT below does not remove the source "Comments" header from that gate, so
# the comment is still read and the derivations still work.
DERIVED_COLUMNS = ["Invoice Date", "Quote Category"]

# Output schema, in exact output order. GTI Comments is intentionally NOT here:
# its free text is distilled into Quote Category + Invoice Date and no longer
# emitted, but out["GTI Comments"] is still computed internally to drive those
# derivations. These are internal field names (the dict keys used everywhere in
# the writer); the displayed header text comes from DISPLAY_HEADERS below.
FINAL_COLUMNS = [
    "Account No.", "Account Name", "Quote", "Est. Date", "Quote Category",
    "Job Name", "Amount", "Product", "SQFT", "Created By", "Internal Note",
    "Invoice", "Invoice Date",
]

# Displayed header text, decoupled from the internal field names above. Used ONLY
# for the header row cells (and therefore the Excel Table column names, which are
# read from those cells). All data-writing branch logic and width lookups stay
# keyed by the internal field name. The two derived date columns get distinct
# labels ("Date" / "Ordered") so the table has no duplicate column names.
DISPLAY_HEADERS = {
    "Account No.":    "Account",
    "Account Name":   "Customer",
    "Quote":          "Quote",
    "Est. Date":      "Date",
    "Quote Category": "Quote Category",
    "Job Name":       "Job Name",
    "Amount":         "Amount",
    "Product":        "Product",
    "SQFT":           "SQFT",
    "Created By":     "Created By",
    "Internal Note":  "Internal Note",
    "Invoice":        "Invoice",
    "Invoice Date":   "Ordered",
}

# 6.8 Created By canonical map.
#
# The mapping is intentionally EMPTY in this public repository so that no real
# operator names are committed. It is supplied at run time by the browser (the
# "Manage names" panel, stored locally per-user) and threaded through
# clean_workbook(..., created_by_map=...).
#
# Shape: { "RAW AS TYPED": "Clean Name", ... }. To leave an already-clean name
# untouched and un-flagged, map it to itself ("Some Name": "Some Name"). A name
# that is neither a key nor one of the clean values is left exactly as-is and
# flagged in the run summary for human review (never guessed).
DEFAULT_CREATED_BY_MAP = {}

# 6.4 field roles.
SINGLE_VALUE_FIELDS = {"Account No.", "Quote", "Est. Date", "Amount", "Invoice"}
SPACE_JOINED_FIELDS = {"Account Name", "Job Name", "GTI Comments",
                       "Created By", "SQFT", "Internal Note"}
# Single-line space-joined fields collapse internal runs of whitespace (PRD 6.4).
# Product is multi-line (split on ';') so it is handled separately.

ID_TEXT_FIELDS = {"Account No.", "Quote", "Invoice"}   # 6.6 keep as text ('@')
NUMERIC_FIELDS = {"Amount", "SQFT"}                    # 6.7 numbers, '#,##0.00'
DATE_FIELDS = {"Est. Date", "Invoice Date"}            # 6.5 real dates, 'mm/dd/yyyy'
WRAP_FIELDS = {"Product", "GTI Comments", "Job Name", "Account Name",
               "Internal Note"}  # 6.11

ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
NUMERIC_QUOTE_RE = re.compile(r"^\d+$")


# ---------------------------------------------------------------------------
# 6.12 GTI Comments / Created By derivations (Invoice Date + Quote Category).
#
# Quote Category used to be derived purely from the shape of the GTI Comments
# text. That broke down for invoiced quotes: GTI Comments holds only the
# LATEST lifecycle event, so once a quote is invoiced its comment becomes
# "Created Invoice Number: N on <date>" and any earlier origin marker (web
# quote / duplicate) is gone for good. Comment shape alone can no longer
# answer "was this quote originally a web quote" once it has been invoiced.
#
# Created By does not have that problem — it is a separate source column that
# invoicing never touches. Every quote whose comment says "Estimate created
# from Web Quote N" has Created By == WEBSITE, whether or not it was later
# invoiced, and WEBSITE never appears on a duplicated or blank-comment quote.
# So Quote Category is now a 3-way split: duplication is checked first from
# the comment (it is defined by the action taken, and duplication is always
# performed by a CSR, never through the website, so it can never collide with
# the WEBSITE check); web origin is then checked from the RAW Created By
# value; everything else — including every blank-comment CSR-handled quote —
# falls to Phone/Email. There is no residual "Other" category: an unrecognized
# comment shape is still flagged (see _is_recognized_comment_shape) but no
# longer changes what category a row gets.
#
# Invoice Date is unrelated to Quote Category: it is derived purely from
# INVOICE_COMMENT_RE matching the comment, independent of what category the
# row ends up in.
# ---------------------------------------------------------------------------

CATEGORY_DUPLICATED = "Duplicated Quote"  # comment matches DUPLICATE_QUOTE_COMMENT_RE
CATEGORY_WEB = "Web Quote"                # else, raw Created By == "WEBSITE"
CATEGORY_PHONE_EMAIL = "Phone/Email"      # else (default; includes blank comments)

CATEGORY_ORDER = [CATEGORY_DUPLICATED, CATEGORY_WEB, CATEGORY_PHONE_EMAIL]

INVOICE_COMMENT_RE = re.compile(r"^Created Invoice Number:\s*(\d+)\s+on\b\s*(.*)$")
WEB_QUOTE_COMMENT_RE = re.compile(r"^Estimate created from Web Quote\s+\d+\s*$")
DUPLICATE_QUOTE_COMMENT_RE = re.compile(r"^Quote based on Est No:\s*\d+\s*$")

# The date fragment following "on" (e.g. "Jun 30 2026 10:25AM"). Invoice Date is
# stored date-only to match the Est. Date column, so the time is parsed but
# discarded. A few comments carry the date with no time at all; those still
# yield a date.
INVOICE_DATE_RE = re.compile(r"^([A-Za-z]{3})\s+(\d{1,2})\s+(\d{4})\b")

# When a record straddles a page break the export can push the date fragment of
# its invoice comment onto the following page's sub-header line, which furniture
# removal then discards. Such a line holds a bare date-time and nothing else in
# the Comments column, which makes the value recoverable rather than lost. See
# _collect_stranded_dates.
STRANDED_DATE_RE = re.compile(
    r"^([A-Za-z]{3})\s+(\d{1,2})\s+(\d{4})\s+\d{1,2}:\d{2}\s*[AaPp][Mm]$")

# Explicit month table rather than strptime('%b'), which is locale-dependent.
MONTH_ABBR = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
              "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


class CleaningError(Exception):
    """Raised to abort cleaning with a human-readable message (validation gates)."""


# ---------------------------------------------------------------------------
# Cell reading — normalize xlrd cells to native Python values, preserving type.
# ---------------------------------------------------------------------------

def _read_grid(raw_bytes):
    """Return a list of rows; each cell normalized to str | float | date | None."""
    book = xlrd.open_workbook(file_contents=raw_bytes)
    sheet = book.sheet_by_index(0)
    datemode = book.datemode
    grid = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            ct = cell.ctype
            if ct == xlrd.XL_CELL_TEXT:
                row.append(cell.value)
            elif ct == xlrd.XL_CELL_NUMBER:
                row.append(float(cell.value))
            elif ct == xlrd.XL_CELL_DATE:
                # Real Excel date cell -> Python date (PRD 6.5 robust path).
                dt = xlrd.xldate_as_datetime(cell.value, datemode)
                row.append(dt.date())
            elif ct == xlrd.XL_CELL_BOOLEAN:
                row.append(bool(cell.value))
            else:  # EMPTY, BLANK, ERROR
                row.append(None)
        grid.append(row)
    return grid


def _text(value):
    """Stripped string form of a normalized cell ('' for blank/None)."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float):
        # Integer-valued floats render without a trailing '.0'.
        if value == int(value):
            return str(int(value))
        return repr(value)
    if isinstance(value, datetime.date):
        return value.isoformat()
    return str(value).strip()


# ---------------------------------------------------------------------------
# Header detection (PRD 6.3) — anchored on header text, never fixed index.
# ---------------------------------------------------------------------------

def _joined_text(row):
    parts = [_text(v) for v in row]
    return " ".join(p for p in parts if p)


# Header label fragments that appear in the Acct-No. / Est-No. columns on the
# two header lines but never in real data (data there is always numeric).
_ACCT_COL_HEADER_TOKENS = {"Acct", "No."}
_QUOTE_COL_HEADER_TOKENS = {"Est", "No."}


def _is_page_marker(row):
    """True for furniture detectable without the column map (Gate 1, PRD 6.1)."""
    joined = _joined_text(row)
    if not joined:
        return False  # fully-blank rows are continuations, not furniture
    for marker in FURNITURE_SUBSTRINGS:
        if marker in joined:
            return True
    if joined.startswith("Acct"):        # the main column-header line
        return True
    tokens = [_text(v) for v in row if _text(v)]
    if tokens and all(t in SUBHEADER_TOKENS for t in tokens):
        return True                       # clean "No." / "By" sub-header line
    return False


def _is_furniture(row, acol=None, qcol=None):
    """True if the row is page furniture (PRD 6.1).

    With the Acct-No. (acol) and Quote (qcol) column indices known, the header
    and sub-header lines are recognized by their header-label fragments even
    when a stray wrapped value (e.g. an invoice timestamp) bleeds into the row.
    """
    if _is_page_marker(row):
        return True
    if acol is not None and _text(row[acol]) in _ACCT_COL_HEADER_TOKENS:
        return True
    if qcol is not None and _text(row[qcol]) in _QUOTE_COL_HEADER_TOKENS:
        return True
    return False


def _build_header_map(grid):
    """Locate the first header block and map source label -> column index.

    The header spans two physical rows (e.g. 'Acct'/'No.' -> 'Acct No.').
    """
    header_row_idx = None
    for i, row in enumerate(grid):
        if _joined_text(row).startswith("Acct"):
            header_row_idx = i
            break
    if header_row_idx is None:
        return None
    top = grid[header_row_idx]
    bottom = grid[header_row_idx + 1] if header_row_idx + 1 < len(grid) else []
    ncols = len(top)
    label_to_col = {}
    for c in range(ncols):
        t = _text(top[c]) if c < len(top) else ""
        b = _text(bottom[c]) if c < len(bottom) else ""
        label = (t + " " + b).strip() if (t and b) else (t or b)
        if label and label not in label_to_col:
            label_to_col[label] = c
    return label_to_col


# ---------------------------------------------------------------------------
# Row consolidation (PRD 6.4) with page-split de-duplication.
# ---------------------------------------------------------------------------

def _join_fragments(fragments, single_line):
    """Space-join fragments, collapsing consecutive identical fragments.

    The consecutive-duplicate collapse handles page-split records where the
    export repeats a single-value field on the continuation row (e.g. a Job Name
    appearing identically on both physical rows). Genuinely different wrap
    fragments (e.g. a company name whose first half ends one page and second
    half begins the next, 'ACME GLASS' + 'PORTLAND, LLC.') are preserved.
    """
    kept = []
    for frag in fragments:
        frag = frag.strip()
        if not frag:
            continue
        if kept and kept[-1] == frag:
            continue
        kept.append(frag)
    joined = " ".join(kept)
    if single_line:
        joined = re.sub(r"\s+", " ", joined).strip()
    return joined


def _resolve_date(cell):
    """(kind, value) for an Est. Date cell. kind in {'date','text','blank'} (PRD 6.5)."""
    if cell is None or (isinstance(cell, str) and cell.strip() == ""):
        return ("blank", None)
    if isinstance(cell, datetime.datetime):
        return ("date", cell.date())
    if isinstance(cell, datetime.date):
        return ("date", cell)
    if isinstance(cell, str) and ISO_DATE_RE.match(cell.strip()):
        try:
            d = datetime.date.fromisoformat(cell.strip()[:10])
            return ("date", d)
        except ValueError:
            return ("text", cell.strip())
    return ("text", _text(cell))


def _consolidate(data_rows, col):
    """Build one record per distinct-adjacent Quote from furniture-free rows (PRD 6.4).

    `col` maps final-column-name -> source column index.
    Returns (records, flags).
    """
    q_idx = col["Quote"]
    records = []
    seen_quotes = {}
    nonadjacent_dupes = []
    active = None

    def new_record(primary_row, quote):
        rec = {"quote": quote, "single": {}, "frag": {}, "product": []}
        for name in SINGLE_VALUE_FIELDS:
            rec["single"][name] = primary_row[col[name]]
        for name in SPACE_JOINED_FIELDS:
            rec["frag"][name] = []
        _accumulate(rec, primary_row)
        return rec

    def _accumulate(rec, row):
        for name in SPACE_JOINED_FIELDS:
            rec["frag"][name].append(_text(row[col[name]]))
        rec["product"].append(_text(row[col["Product"]]))

    for row in data_rows:
        quote = _text(row[q_idx])
        if quote != "" and (active is None or quote != active["quote"]):
            if quote in seen_quotes:
                nonadjacent_dupes.append(quote)
            active = new_record(row, quote)
            records.append(active)
            seen_quotes[quote] = len(records) - 1
        else:
            if active is None:
                continue  # leading blank/continuation before any quote: nothing to attach
            _accumulate(active, row)

    return records, {"nonadjacent_dupes": nonadjacent_dupes}


# ---------------------------------------------------------------------------
# GTI Comments derivations (PRD 6.12)
# ---------------------------------------------------------------------------

def _month_day_year(mon, day, year):
    """A real date from an abbreviated month / day / year, or None if invalid."""
    month = MONTH_ABBR.get(mon.strip().lower())
    if month is None:
        return None
    try:
        return datetime.date(int(year), month, int(day))
    except ValueError:
        return None


def _parse_comment_date(fragment):
    """Date from a comment date fragment, or None if there isn't a real one.

    'Jun 30 2026 10:25AM' and 'Jun 30 2026' both give date(2026, 6, 30); the
    time of day is deliberately dropped. '' (a comment truncated at "on") and
    anything unparseable give None.
    """
    if not fragment:
        return None
    m = INVOICE_DATE_RE.match(fragment.strip())
    if not m:
        return None
    return _month_day_year(m.group(1), m.group(2), m.group(3))


def _derive_quote_category(comment, created_by_raw):
    """Quote Category for one finalized record (PRD 6.12, Created-By-aware).

    Priority: Duplicated Quote (comment shape) first and unconditionally, then
    Web Quote (raw, pre-mapping Created By == "WEBSITE"), then Phone/Email as
    the catch-all. created_by_raw must be the joined Created By fragment BEFORE
    _map_created_by() runs, so a user's optional name-normalization mapping can
    never relabel a web-origin quote by remapping the literal string "WEBSITE".
    """
    text = comment.strip()
    if DUPLICATE_QUOTE_COMMENT_RE.match(text):
        return CATEGORY_DUPLICATED
    if created_by_raw.strip().upper() == "WEBSITE":
        return CATEGORY_WEB
    return CATEGORY_PHONE_EMAIL


def _invoice_date_fragment(comment):
    """Raw text following "on" in an invoice comment, or None if it isn't one.

    Independent of Quote Category: fires purely off INVOICE_COMMENT_RE against
    the comment text. The fragment may be '' when the export truncated the
    comment right at "on".
    """
    m = INVOICE_COMMENT_RE.match(comment.strip())
    if m:
        return m.group(2).strip()
    return None


def _is_recognized_comment_shape(comment):
    """True for a blank comment or one of the 3 known non-blank shapes.

    Purely a data-quality safety net (flags["unrecognized_comments"]),
    decoupled from Quote Category — format drift should surface here even
    though WEB_QUOTE_COMMENT_RE no longer drives a category on its own.
    """
    text = comment.strip()
    if text == "":
        return True
    return bool(
        INVOICE_COMMENT_RE.match(text)
        or WEB_QUOTE_COMMENT_RE.match(text)
        or DUPLICATE_QUOTE_COMMENT_RE.match(text)
    )


def _collect_stranded_dates(grid, acol, qcol, ccol):
    """Map quote -> date for invoice dates orphaned on a page sub-header line.

    Scanned over the full grid, before furniture removal, because that is where
    the fragment still exists. A furniture line whose Comments cell holds a bare
    date-time and nothing else is the tail of the invoice comment belonging to
    the record that was in progress when the page broke.

    The result is only consulted when that record's own comment ends at "on"
    with no date of its own (see _finalize_record), so a value found here can
    never displace a date the comment already carries.
    """
    stranded = {}
    active = None
    for row in grid:
        if _is_furniture(row, acol, qcol):
            fragment = _text(row[ccol]) if ccol < len(row) else ""
            fragment = re.sub(r"\s+", " ", fragment).strip()
            m = STRANDED_DATE_RE.match(fragment)
            if m and active:
                d = _month_day_year(m.group(1), m.group(2), m.group(3))
                if d is not None:
                    stranded[active] = d
            continue
        quote = _text(row[qcol])
        if quote:
            active = quote
    return stranded


# ---------------------------------------------------------------------------
# Field finalization
# ---------------------------------------------------------------------------

def _finalize_record(rec, flags, cb_map, cb_known, stranded):
    """Produce the output values for one record.

    stranded: the quote -> date map from _collect_stranded_dates, consulted only
    for an invoice comment that carries no date of its own.
    """
    out = {}

    # Single-value identity/date/number fields (from the primary row).
    out["Account No."] = _text(rec["single"]["Account No."])
    out["Quote"] = _text(rec["single"]["Quote"])
    out["Invoice"] = _text(rec["single"]["Invoice"])

    kind, dval = _resolve_date(rec["single"]["Est. Date"])
    if kind == "text":
        flags["unparseable_dates"].append(out["Quote"])
    out["Est. Date"] = (kind, dval)

    out["Amount"] = _to_number(rec["single"]["Amount"])

    # Space-joined single-line fields.
    for name in ("Account Name", "Job Name", "GTI Comments", "Internal Note"):
        out[name] = _join_fragments(rec["frag"][name], single_line=True)

    # SQFT: space-joined then parsed back to a number (6.4 + 6.7).
    sqft_text = _join_fragments(rec["frag"]["SQFT"], single_line=True)
    out["SQFT"] = _sqft_to_number(sqft_text)

    # Created By: joined, then canonical map (6.8); unknown left as-is + flagged.
    cb = _join_fragments(rec["frag"]["Created By"], single_line=True)
    out["Created By"] = _map_created_by(cb, flags, cb_map, cb_known)

    # Product: joined then split on ';' to newline-separated specs (6.4).
    out["Product"] = _finalize_product(rec["product"])

    # Quote Category (6.12): Duplicated Quote > Web Quote (raw Created By) >
    # Phone/Email. Uses `cb`, the pre-mapping raw value, not out["Created By"].
    out["Quote Category"] = _derive_quote_category(out["GTI Comments"], cb)

    # Invoice Date (6.12): fully independent of Quote Category, fired purely
    # off the comment text matching INVOICE_COMMENT_RE.
    date_fragment = _invoice_date_fragment(out["GTI Comments"])
    invoice_date = None
    if date_fragment is not None:
        invoice_date = _parse_comment_date(date_fragment)
        if invoice_date is None:
            invoice_date = stranded.get(out["Quote"])
            if invoice_date is not None:
                flags["recovered_invoice_dates"].append(out["Quote"])
        if invoice_date is None:
            flags["invoice_without_date"].append(out["Quote"])

    # Data-quality safety net, decoupled from Quote Category.
    if not _is_recognized_comment_shape(out["GTI Comments"]):
        flags["unrecognized_comments"].append(out["GTI Comments"])

    out["Invoice Date"] = ("date", invoice_date) if invoice_date else ("blank", None)

    return out


def _to_number(cell):
    if isinstance(cell, float):
        return cell
    if isinstance(cell, str):
        s = cell.strip().replace(",", "")
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _sqft_to_number(text):
    if text == "":
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None  # doubled/garbled SQFT -> leave empty rather than guess


def _map_created_by(value, flags, cb_map, cb_known):
    if value == "":
        return ""
    if value in cb_map:
        return cb_map[value]
    if value in cb_known:
        return value
    flags["unknown_created_by"].add(value)
    return value


def _finalize_product(fragments):
    joined = _join_fragments(fragments, single_line=True)
    if joined == "":
        return ""
    specs = [p.strip() for p in joined.split(";")]
    specs = [p for p in specs if p]
    return "\n".join(specs)


# ---------------------------------------------------------------------------
# Workbook writing (PRD 6.11)
# ---------------------------------------------------------------------------

# Every column in FINAL_COLUMNS needs an entry here; the lookup is per-column
# and a missing key is a KeyError, not a silent default.
COLUMN_WIDTHS = {
    "Account No.": 10, "Account Name": 26, "Quote": 12, "Est. Date": 12,
    "Job Name": 20, "Amount": 12, "Invoice": 12, "GTI Comments": 28,
    "Product": 34, "SQFT": 10, "Created By": 12, "Internal Note": 34,
    "Invoice Date": 13, "Quote Category": 16,
}


def _write_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"

    arial = "Arial"
    header_font = Font(name=arial, bold=True)
    body_font = Font(name=arial)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    ws.append([DISPLAY_HEADERS[name] for name in FINAL_COLUMNS])
    for c, name in enumerate(FINAL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font

    for out in rows:
        r = ws.max_row + 1
        for c, name in enumerate(FINAL_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c)
            value = out[name]
            cell.font = body_font
            cell.alignment = wrap if name in WRAP_FIELDS else top

            if name in ID_TEXT_FIELDS:
                cell.value = value if value != "" else None
                cell.number_format = "@"
            elif name in DATE_FIELDS:
                kind, dval = value
                if kind == "date":
                    cell.value = dval
                    cell.number_format = "mm/dd/yyyy"
                elif kind == "text":
                    cell.value = dval
                    cell.number_format = "@"
                else:
                    cell.value = None
            elif name in NUMERIC_FIELDS:
                cell.value = value
                cell.number_format = "#,##0.00"
            else:
                cell.value = value if value != "" else None

    # Column widths.
    for c, name in enumerate(FINAL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = COLUMN_WIDTHS[name]

    # Excel Table: banded rows + filter header (PRD 6.11).
    last_row = ws.max_row
    last_col = get_column_letter(len(FINAL_COLUMNS))
    ref = "A1:%s%d" % (last_col, last_row)
    table = Table(displayName="Quotes", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"  # frozen header row

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Summary + gates
# ---------------------------------------------------------------------------

def build_summary(raw_row_count, data_row_count, empty_cols_dropped,
                  records, rows, flags):
    invoiced = sum(1 for out in rows if out["Invoice"] != "")
    total_amount = sum((out["Amount"] or 0.0) for out in rows)

    # Category counts, emitted in a fixed order so the summary is stable. Every
    # row now gets one of the 3 real category values (no blank/residual bucket).
    counts = {c: 0 for c in CATEGORY_ORDER}
    for out in rows:
        counts[out["Quote Category"]] += 1
    categories = {c: counts[c] for c in CATEGORY_ORDER}

    invoice_dates = sum(1 for out in rows if out["Invoice Date"][0] == "date")

    return {
        "rows_in": raw_row_count,
        "rows_after_furniture": data_row_count,
        "empty_columns_dropped": empty_cols_dropped,
        "quotes_out": len(records),
        "quotes_ordered": invoiced,
        "total_amount": round(total_amount, 2),
        "comment_categories": categories,
        "invoice_dates_extracted": invoice_dates,
        "invoice_dates_recovered_from_page_break": len(flags["recovered_invoice_dates"]),
        "flag_unparseable_dates": sorted(set(flags["unparseable_dates"])),
        "flag_unknown_created_by": sorted(flags["unknown_created_by"]),
        "flag_nonadjacent_duplicate_quotes": sorted(set(flags["nonadjacent_dupes"])),
        "flag_invoice_comments_without_date": sorted(set(flags["invoice_without_date"])),
        "flag_unrecognized_comments": sorted(set(flags["unrecognized_comments"]))[:20],
    }


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def clean_workbook(raw_bytes, created_by_map=None):
    """Clean a raw .xls export.

    created_by_map: optional { "RAW": "Clean" } mapping supplied at run time
    (no operator names are stored in this repository). Unknown names are left
    as-is and flagged.
    """
    try:
        return _clean(raw_bytes, created_by_map)
    except CleaningError as exc:
        return {"ok": False, "error": str(exc), "xlsx": None, "summary": None}


def _clean(raw_bytes, created_by_map=None):
    cb_map = dict(created_by_map) if created_by_map else dict(DEFAULT_CREATED_BY_MAP)
    # A name is "known" (and never flagged) if it is either a raw key or one of
    # the clean target values in the supplied map.
    cb_known = set(cb_map.keys()) | set(cb_map.values())

    grid = _read_grid(raw_bytes)
    raw_row_count = len(grid)

    # Gate 1 (PRD 9.1): furniture markers present -> confirms report type.
    if not any(_is_page_marker(row) for row in grid):
        raise CleaningError(
            "This file does not look like a GTI Quotes List export: none of the "
            "expected page markers (QUOTES LIST / FOR DATE RANGE / column headers) "
            "were found. No file was produced."
        )

    # Header map (PRD 6.3) + Gate 2 (PRD 9.2): all required headers mapped.
    label_to_col = _build_header_map(grid)
    if not label_to_col:
        raise CleaningError(
            "Could not locate the column-header row in this export. No file was produced."
        )
    missing = [s for s in REQUIRED_SOURCE_HEADERS if s not in label_to_col]
    if missing:
        raise CleaningError(
            "This export is missing expected column(s): %s. The report format may "
            "have changed. No file was produced." % ", ".join(missing)
        )
    col = {final: label_to_col[source] for final, source in FINAL_TO_SOURCE}
    acol, qcol = col["Account No."], col["Quote"]

    # Furniture removal (PRD 6.1), header-column-aware so stray wrapped fragments
    # on header lines don't leak through as data.
    data_rows = [row for row in grid if not _is_furniture(row, acol, qcol)]
    data_row_count = len(data_rows)

    # 6.2 empty-column removal (reported; header-anchored mapping already ignores them).
    ncols = max((len(r) for r in grid), default=0)
    empty_cols_dropped = 0
    for c in range(ncols):
        if all(_text(row[c]) == "" for row in data_rows if c < len(row)):
            empty_cols_dropped += 1

    # Invoice-date fragments orphaned on page sub-header lines (6.12). Collected
    # from the full grid because furniture removal has already dropped them from
    # data_rows.
    stranded = _collect_stranded_dates(grid, acol, qcol, col["GTI Comments"])

    records, cflags = _consolidate(data_rows, col)
    flags = {
        "unparseable_dates": [],
        "unknown_created_by": set(),
        "nonadjacent_dupes": cflags["nonadjacent_dupes"],
        "recovered_invoice_dates": [],
        "invoice_without_date": [],
        "unrecognized_comments": [],
    }
    rows = [_finalize_record(rec, flags, cb_map, cb_known, stranded)
            for rec in records]

    # Gate 3 (PRD 9.3): output row count == distinct-adjacent Quote count.
    distinct_adjacent = _count_distinct_adjacent(data_rows, col["Quote"])
    if len(rows) != distinct_adjacent:
        raise CleaningError(
            "Internal consistency check failed: consolidated %d rows but counted %d "
            "distinct quotes. No file was produced." % (len(rows), distinct_adjacent)
        )

    # Gate 5 (PRD 9.5): every output row has a numeric-string Quote.
    bad_quotes = [out["Quote"] for out in rows
                  if not NUMERIC_QUOTE_RE.match(out["Quote"] or "")]
    if bad_quotes:
        raise CleaningError(
            "Some consolidated records have a non-numeric Quote value (%s). No file "
            "was produced." % ", ".join(bad_quotes[:5])
        )

    xlsx = _write_workbook(rows)
    summary = build_summary(raw_row_count, data_row_count, empty_cols_dropped,
                            records, rows, flags)
    return {"ok": True, "error": None, "xlsx": xlsx, "summary": summary}


def _count_distinct_adjacent(data_rows, q_idx):
    """Count of rows whose Quote differs from the current active quote (PRD 6.4)."""
    count = 0
    active = None
    for row in data_rows:
        quote = _text(row[q_idx])
        if quote != "" and (active is None or quote != active):
            count += 1
            active = quote
    return count
